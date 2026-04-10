"""
rag_pipeline.py — Real RAG Pipeline for FallahTech Investment Analysis
=======================================================================
Architecture:
  1. INGESTION   : Load PDFs (PyMuPDF) + Excel (openpyxl) → raw Documents
  2. CHUNKING    : RecursiveCharacterTextSplitter (800 tokens, overlap 100)
  3. EMBEDDING   : HuggingFace all-MiniLM-L6-v2 → dense vectors
  4. INDEXING    : FAISS vector store (persisted to disk)
  5. RETRIEVAL   : Similarity search, top-k=4 chunks
  6. GENERATION  : Mistral via Ollama with retrieved context + strict prompt
"""

import warnings
warnings.filterwarnings("ignore")

import os
import requests
import fitz  # PyMuPDF
import pickle

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
from openpyxl import load_workbook

# ── Config ────────────────────────────────────────────────────────────────────
OLLAMA_URL    = "http://localhost:11434/api/generate"
OLLAMA_MODEL  = "mistral"
EMBEDDING_MODEL = "all-MiniLM-L6-v2"
CHUNK_SIZE    = 800
CHUNK_OVERLAP = 100
TOP_K         = 4
FAISS_INDEX   = "faiss_index"   # folder where index is saved on disk

# Task 3 scoring grid
CRITERIA = {
    "Financier":  40,
    "Commercial": 30,
    "Équipe":     15,
    "Marché":     15,
}


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — INGESTION
# ═══════════════════════════════════════════════════════════════════════════════

def load_pdf(pdf_path: str) -> list[Document]:
    """Load a single PDF with PyMuPDF — handles custom fonts better than PyPDFLoader."""
    docs = []
    try:
        pdf = fitz.open(pdf_path)
        for page_num, page in enumerate(pdf):
            text = page.get_text().strip()
            if text:
                docs.append(Document(
                    page_content=text,
                    metadata={
                        "source":    os.path.basename(pdf_path),
                        "page":      page_num + 1,
                        "file_type": "pdf"
                    }
                ))
        pdf.close()
        print(f"  [PDF] {os.path.basename(pdf_path)} — {len(docs)} pages")
    except Exception as e:
        print(f"  [WARN] Cannot read {pdf_path}: {e}")
    return docs


def load_all_pdfs(folder: str) -> list[Document]:
    """Load every PDF in a folder."""
    docs = []
    for fname in sorted(os.listdir(folder)):
        if fname.lower().endswith(".pdf"):
            docs.extend(load_pdf(os.path.join(folder, fname)))
    return docs


def load_excel(xlsx_path: str) -> list[Document]:
    """
    Load Excel file sheet by sheet.
    Each sheet becomes ONE document so row relationships are preserved.
    """
    docs = []
    wb = load_workbook(xlsx_path, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not any(v is not None for v in row):
                continue
            row_str = " | ".join(str(v) if v is not None else "" for v in row)
            if i == 0:
                headers = row_str  # keep header for context
            rows.append(row_str)

        content = (
            f"[FICHIER: {os.path.basename(xlsx_path)}]\n"
            f"[FEUILLE: {sheet_name}]\n"
            f"[COLONNES: {headers}]\n\n"
            + "\n".join(rows)
        )
        docs.append(Document(
            page_content=content,
            metadata={
                "source":    os.path.basename(xlsx_path),
                "sheet":     sheet_name,
                "file_type": "excel"
            }
        ))
    print(f"  [XLS] {os.path.basename(xlsx_path)} — {len(docs)} feuilles")
    return docs


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2+3+4 — CHUNKING + EMBEDDING + INDEXING
# ═══════════════════════════════════════════════════════════════════════════════

def build_vectorstore(documents: list[Document], save_path: str) -> FAISS:
    """Chunk → embed → index → save to disk."""

    # CHUNKING
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
        separators=["\n[FEUILLE", "\n[FICHIER", "\n\n", "\n", ". ", " "],
        keep_separator=True
    )
    chunks = splitter.split_documents(documents)
    print(f"\n  [CHUNK] {len(documents)} docs → {len(chunks)} chunks "
          f"(size={CHUNK_SIZE}, overlap={CHUNK_OVERLAP})")

    # EMBEDDING
    print(f"  [EMBED] model={EMBEDDING_MODEL} ...")
    embeddings = HuggingFaceEmbeddings(
        model_name=EMBEDDING_MODEL,
        model_kwargs={"device": "cpu"},
        encode_kwargs={"normalize_embeddings": True}   # cosine similarity
    )

    # INDEXING — FAISS
    print("  [FAISS] Building index ...")
    vectorstore = FAISS.from_documents(chunks, embeddings)

    # PERSIST to disk so we don't re-embed every restart
    vectorstore.save_local(save_path)
    print(f"  [FAISS] Index saved → {save_path}/")

    return vectorstore


def load_vectorstore(save_path: str) -> FAISS:
    """Load persisted FAISS index from disk."""
    embeddings = HuggingFaceEmbeddings(
        model_name=EMBEDDING_MODEL,
        model_kwargs={"device": "cpu"},
        encode_kwargs={"normalize_embeddings": True}
    )
    return FAISS.load_local(
        save_path, embeddings,
        allow_dangerous_deserialization=True
    )


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5 — RETRIEVAL
# ═══════════════════════════════════════════════════════════════════════════════

def retrieve(vectorstore: FAISS, query: str, k: int = TOP_K) -> list[Document]:
    """
    Semantic similarity search.
    Returns top-k chunks most relevant to the query.
    """
    return vectorstore.similarity_search(query, k=k)


def format_context(docs: list[Document]) -> str:
    """Format retrieved chunks into a readable context block."""
    parts = []
    for i, doc in enumerate(docs, 1):
        source = doc.metadata.get("source", "?")
        page   = doc.metadata.get("page", "")
        sheet  = doc.metadata.get("sheet", "")
        loc    = f"page {page}" if page else f"feuille '{sheet}'" if sheet else ""
        parts.append(
            f"[EXTRAIT {i} — {source}{', ' + loc if loc else ''}]\n"
            f"{doc.page_content.strip()}"
        )
    return "\n\n" + "─" * 60 + "\n\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6 — GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def _call_ollama(prompt: str, max_tokens: int = 500) -> str:
    """Call Ollama Mistral with safe context limit."""
    # Trim to stay within num_ctx
    if len(prompt) > 3800:
        prompt = prompt[:3800] + "\n...[contexte tronqué pour respecter la limite]"
    try:
        resp = requests.post(
            OLLAMA_URL,
            json={
                "model":  OLLAMA_MODEL,
                "prompt": prompt,
                "stream": False,
                "options": {
                    "temperature": 0.1,    # low = factual, high = creative
                    "num_predict": max_tokens,
                    "num_ctx":     2048,   # context window
                    "num_thread":  4,
                }
            },
            timeout=180
        )
        resp.raise_for_status()
        return resp.json().get("response", "Pas de réponse.")
    except requests.exceptions.HTTPError as e:
        return f"[Erreur Ollama {e.response.status_code}] Redémarrez Ollama."
    except requests.exceptions.ConnectionError:
        return "[Erreur] Ollama hors ligne. Lancez : ollama serve"
    except requests.exceptions.Timeout:
        return "[Erreur] Timeout Ollama. Réessayez."
    except Exception as e:
        return f"[Erreur] {e}"


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN RAG SYSTEM CLASS
# ═══════════════════════════════════════════════════════════════════════════════

class RAGSystem:
    """
    Full RAG pipeline:
      - Ingests PDFs + Excel on first run, persists FAISS index
      - On subsequent runs loads index from disk (fast startup)
      - Supports free questions and structured investment scoring
    """

    def __init__(self, pdf_folder: str, xlsx_path: str = None,
                 force_rebuild: bool = False):

        self.pdf_folder = pdf_folder
        self.xlsx_path  = xlsx_path

        base_dir   = os.path.dirname(os.path.abspath(__file__))
        index_path = os.path.join(base_dir, FAISS_INDEX)

        # Load existing index OR build from scratch
        if os.path.exists(index_path) and not force_rebuild:
            print(f"\n✅ Loading existing FAISS index from {index_path} ...")
            self.vectorstore = load_vectorstore(index_path)
        else:
            print("\n🔄 Building FAISS index from documents ...")
            documents = []
            documents.extend(load_all_pdfs(pdf_folder))
            if xlsx_path and os.path.exists(xlsx_path):
                documents.extend(load_excel(xlsx_path))
            print(f"\n  Total raw documents : {len(documents)}")
            self.vectorstore = build_vectorstore(documents, index_path)

        print("✅ RAG system ready.\n")

    # ── Free question (works for ANY question) ────────────────────────────────
    def query(self, question: str) -> str:
        """
        Full RAG pipeline for any question:
          1. Embed question
          2. Retrieve top-4 relevant chunks
          3. Build prompt with context
          4. Generate answer with Mistral
          5. Return answer + sources + investment signal
        """
        # RETRIEVAL
        retrieved_docs = retrieve(self.vectorstore, question, k=TOP_K)
        context        = format_context(retrieved_docs)
        sources_list   = list({
            f"{d.metadata.get('source')} "
            f"{'p.' + str(d.metadata.get('page')) if d.metadata.get('page') else ''}"
            for d in retrieved_docs
        })

        # GENERATION
        prompt = f"""Tu es un analyste financier expert travaillant pour un fonds d'investissement.
Tu évalues le dossier de FallahTech SARL en vue d'une décision d'investissement Série A.

RÈGLES STRICTES :
1. Réponds UNIQUEMENT à partir des extraits fournis ci-dessous
2. Cite toujours la source exacte (fichier + page ou feuille)
3. Si l'information est absente → écris : "Information non disponible dans le corpus."
4. NE JAMAIS inventer, calculer ou extrapoler un chiffre absent
5. Termine par un SIGNAL D'INVESTISSEMENT lié à cette question

EXTRAITS RÉCUPÉRÉS (top-{TOP_K} par similarité sémantique) :
{context}

QUESTION DE L'ANALYSTE : {question}

FORMAT DE RÉPONSE OBLIGATOIRE :
RÉPONSE: [réponse factuelle avec chiffres exacts et source]
SOURCE: [fichier:page ou fichier:feuille]
SIGNAL D'INVESTISSEMENT: [✅ Positif / ⚠️ Neutre / ❌ Négatif] — [explication en 1 phrase pour le comité]"""

        raw = _call_ollama(prompt, max_tokens=500)

        # Parse structured output
        reponse = signal = source = ""
        for line in raw.splitlines():
            if line.startswith("RÉPONSE:"):
                reponse = line.replace("RÉPONSE:", "").strip()
            elif line.startswith("SOURCE:"):
                source  = line.replace("SOURCE:", "").strip()
            elif line.startswith("SIGNAL D'INVESTISSEMENT:"):
                signal  = line.replace("SIGNAL D'INVESTISSEMENT:", "").strip()

        # Fallback: model didn't follow format → show raw answer
        if not reponse:
            reponse = raw

        # Build final output
        output = reponse
        if source:
            output += f"\n\n📄 **Source :** {source}"
        elif sources_list:
            output += f"\n\n📄 **Sources consultées :** {', '.join(sources_list)}"
        if signal:
            output += f"\n\n🎯 **Signal d'investissement :** {signal}"

        return output

    # ── Investment scoring (Task 3) ───────────────────────────────────────────
    def _score_criterion(self, criterion: str, weight: int) -> dict:
        """Score one criterion using RAG retrieval."""

        # Craft a targeted retrieval query per criterion
        queries = {
            "Financier":  f"FallahTech chiffre affaires résultat net marge bilan trésorerie ratio financier",
            "Commercial": f"FallahTech clients revenus modèle commercial ARPU taux rétention croissance ventes",
            "Équipe":     f"FallahTech fondateurs équipe expérience compétences recrutement",
            "Marché":     f"FallahTech marché taille AgriTech Tunisie concurrents différenciation positionnement",
        }
        query = queries.get(criterion, f"FallahTech {criterion}")

        # RETRIEVAL
        retrieved_docs = retrieve(self.vectorstore, query, k=TOP_K)
        context        = format_context(retrieved_docs)

        # GENERATION
        prompt = f"""Tu es un analyste financier expert évaluant FallahTech SARL pour un investissement Série A.

CRITÈRE À ÉVALUER : {criterion} (pondération dans la grille : {weight}%)

EXTRAITS RÉCUPÉRÉS DES DOCUMENTS (top-{TOP_K} par similarité) :
{context}

RÈGLES :
- Utilise UNIQUEMENT les données des extraits ci-dessus
- NE CALCULE PAS et NE DÉDUIS PAS de valeurs absentes
- Si absent → "Non disponible dans le corpus"

Réponds EXACTEMENT dans ce format :
SCORE: X/10
SOURCE: [fichier exact + page ou feuille]
JUSTIFICATION: [1 phrase avec les chiffres lus directement dans les extraits]"""

        text = _call_ollama(prompt, max_tokens=250)

        # Parse
        score = 5.0
        source = "inconnu"
        justification = text

        for line in text.splitlines():
            if line.startswith("SCORE:"):
                try:
                    score = float(line.replace("SCORE:", "").strip().split("/")[0])
                except Exception:
                    score = 5.0
            elif line.startswith("SOURCE:"):
                source = line.replace("SOURCE:", "").strip()
            elif line.startswith("JUSTIFICATION:"):
                justification = line.replace("JUSTIFICATION:", "").strip()

        return {
            "criterion":      criterion,
            "weight":         weight,
            "score_10":       round(score, 1),
            "weighted_score": round((score / 10) * weight, 2),
            "source":         source,
            "justification":  justification,
        }

    def score_investment(self) -> dict:
        """Run full T3 scoring across all 4 criteria."""
        results = []
        total   = 0
        for criterion, weight in CRITERIA.items():
            print(f"  Scoring {criterion} ({weight}%) ...")
            result = self._score_criterion(criterion, weight)
            results.append(result)
            total += result["weighted_score"]

        if total >= 75:
            decision = "✅ Investir"
        elif total >= 50:
            decision = "⚠️ Investir sous conditions"
        else:
            decision = "❌ No-Go"

        return {
            "criteria":    results,
            "total_score": round(total, 1),
            "decision":    decision,
        }

    def rebuild_index(self):
        """Force rebuild the FAISS index (call if you add new documents)."""
        base_dir   = os.path.dirname(os.path.abspath(__file__))
        index_path = os.path.join(base_dir, FAISS_INDEX)
        documents  = []
        documents.extend(load_all_pdfs(self.pdf_folder))
        if self.xlsx_path and os.path.exists(self.xlsx_path):
            documents.extend(load_excel(self.xlsx_path))
        self.vectorstore = build_vectorstore(documents, index_path)
        print("✅ Index rebuilt.")
